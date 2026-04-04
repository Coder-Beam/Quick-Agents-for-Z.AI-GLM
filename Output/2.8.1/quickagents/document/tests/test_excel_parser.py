"""
Tests for Excel Parser (Phase 5).

Covers:
- T023: Cell data reading (multi-sheet, values, types)
- T024: Formula parsing + dependency tracking
- T025: Chart data extraction (openpyxl chart support)
- T026: Requirement matrix / feature list structure recognition
- T027: Edge cases and regression
"""

import pytest
from pathlib import Path

from quickagents.document.parsers.excel_parser import ExcelParser
from quickagents.document.models import DocumentResult, DocumentFormula

FIXTURES_DIR = Path(__file__).parent / "fixtures"
SAMPLE_XLSX = FIXTURES_DIR / "test_sample.xlsx"


@pytest.fixture
def parser():
    return ExcelParser()


@pytest.fixture
def sample_result(parser):
    return parser.parse(SAMPLE_XLSX)


# ============================================================
# T023: Excel Parser Init + Cell Data Reading
# ============================================================

class TestExcelParserInit:
    def test_availability(self, parser):
        assert parser.is_available()

    def test_parser_name(self, parser):
        assert parser.PARSER_NAME == "excel"

    def test_supported_formats(self, parser):
        assert "xlsx" in parser.SUPPORTED_FORMATS
        assert "xls" in parser.SUPPORTED_FORMATS


class TestExcelCellData:
    def test_is_document_result(self, sample_result):
        assert isinstance(sample_result, DocumentResult)

    def test_source_format(self, sample_result):
        assert sample_result.source_format == "xlsx"

    def test_title(self, sample_result):
        assert sample_result.title == "Requirements"

    def test_no_errors(self, sample_result):
        assert sample_result.errors == []

    def test_sections_for_sheets(self, sample_result):
        assert len(sample_result.sections) == 3
        titles = [s.title for s in sample_result.sections]
        assert "Requirements" in titles
        assert "Calculations" in titles
        assert "Notes" in titles

    def test_section_ids_unique(self, sample_result):
        ids = [s.section_id for s in sample_result.sections]
        assert len(ids) == len(set(ids))

    def test_section_level_all_one(self, sample_result):
        for s in sample_result.sections:
            assert s.level == 1

    def test_section_content_populated(self, sample_result):
        for s in sample_result.sections:
            assert s.content is not None

    def test_raw_text_not_empty(self, sample_result):
        assert len(sample_result.raw_text) > 0

    def test_raw_text_contains_sheet_names(self, sample_result):
        assert "Requirements" in sample_result.raw_text
        assert "Calculations" in sample_result.raw_text

    def test_raw_text_contains_data(self, sample_result):
        assert "REQ-001" in sample_result.raw_text
        assert "Widget A" in sample_result.raw_text

    def test_paragraphs_extracted(self, sample_result):
        assert len(sample_result.paragraphs) > 0

    def test_paragraphs_are_strings(self, sample_result):
        for p in sample_result.paragraphs:
            assert isinstance(p, str)
            assert len(p) > 5

    def test_metadata_present(self, sample_result):
        m = sample_result.metadata
        assert "sheet_count" in m
        assert m["sheet_count"] == 3
        assert "sheet_names" in m
        assert len(m["sheet_names"]) == 3

    def test_metadata_has_creator(self, sample_result):
        assert "creator" in sample_result.metadata
        assert sample_result.metadata["creator"] == "openpyxl"

    def test_structure_tree(self, sample_result):
        tree = sample_result.structure_tree
        assert "children" in tree
        assert len(tree["children"]) == 3


# ============================================================
# T024: Formula Parsing
# ============================================================

class TestExcelFormulas:
    def test_formulas_found(self, sample_result):
        assert len(sample_result.formulas) >= 6

    def test_formula_structure(self, sample_result):
        f = sample_result.formulas[0]
        assert isinstance(f, DocumentFormula)
        assert f.formula_id.startswith("F")
        assert f.formula_text.startswith("=")
        assert f.cell_ref is not None
        assert f.sheet_name == "Calculations"

    def test_formula_dependencies(self, sample_result):
        mul = [f for f in sample_result.formulas if "*" in f.formula_text]
        assert len(mul) >= 3
        for f in mul:
            assert len(f.dependencies) >= 2

    def test_sum_formula(self, sample_result):
        sums = [
            f for f in sample_result.formulas
            if f.formula_text.upper().startswith("=SUM")
        ]
        assert len(sums) >= 1
        assert any("D2:D4" in d for s in sums for d in s.dependencies)

    def test_average_formula(self, sample_result):
        avgs = [
            f for f in sample_result.formulas
            if f.formula_text.upper().startswith("=AVERAGE")
        ]
        assert len(avgs) >= 1

    def test_count_formula(self, sample_result):
        counts = [
            f for f in sample_result.formulas
            if f.formula_text.upper().startswith("=COUNT")
        ]
        assert len(counts) >= 1

    def test_formula_description_not_empty(self, sample_result):
        for f in sample_result.formulas:
            assert f.description, f"Formula {f.formula_id} has no description"

    def test_formula_cell_ref_includes_sheet(self, sample_result):
        for f in sample_result.formulas:
            assert "!" in f.cell_ref, f"Formula {f.formula_id} missing sheet in ref"

    def test_formula_ids_unique(self, sample_result):
        ids = [f.formula_id for f in sample_result.formulas]
        assert len(ids) == len(set(ids))


# ============================================================
# T026: Requirement Matrix Recognition
# ============================================================

class TestRequirementMatrix:
    def test_tables_found(self, sample_result):
        assert len(sample_result.tables) >= 1

    def test_req_matrix_table_has_caption(self, sample_result):
        req_table = sample_result.tables[0]
        assert req_table.caption is not None
        assert "Requirement" in req_table.caption or "matrix" in req_table.caption.lower()

    def test_req_matrix_headers(self, sample_result):
        req_table = sample_result.tables[0]
        headers_lower = [h.lower() for h in req_table.headers]
        assert any("req" in h or "id" in h or "编号" in h for h in headers_lower)

    def test_req_matrix_data_rows(self, sample_result):
        req_table = sample_result.tables[0]
        assert len(req_table.rows) >= 5

    def test_req_matrix_contains_ids(self, sample_result):
        req_table = sample_result.tables[0]
        first_col = [row[0] for row in req_table.rows if row]
        assert any("REQ" in v for v in first_col)

    def test_calculation_table_found(self, sample_result):
        calc_tables = [
            t for t in sample_result.tables
            if t.table_id == "T002"
        ]
        assert len(calc_tables) == 1
        headers_lower = [h.lower() for h in calc_tables[0].headers]
        assert any("item" in h for h in headers_lower)


# ============================================================
# T025: Chart Data (basic openpyxl chart support)
# ============================================================

class TestExcelCharts:
    def test_no_images_for_xlsx(self, sample_result):
        assert sample_result.images == []

    def test_no_crash_on_no_charts(self, sample_result):
        assert isinstance(sample_result, DocumentResult)


# ============================================================
# T027: Edge Cases + Roundtrip
# ============================================================

class TestExcelEdgeCases:
    def test_file_not_found(self, parser):
        with pytest.raises(FileNotFoundError):
            parser.parse(Path("nonexistent.xlsx"))

    def test_to_dict_roundtrip(self, sample_result):
        d = sample_result.to_dict()
        assert d["source_format"] == "xlsx"
        assert len(d["sections"]) == 3
        assert len(d["formulas"]) >= 6
        restored = DocumentResult.from_dict(d)
        assert restored.source_file == sample_result.source_file
        assert len(restored.sections) == len(sample_result.sections)

    def test_get_hash(self, sample_result):
        h = sample_result.get_hash()
        assert isinstance(h, str)
        assert len(h) == 16

    def test_has_errors(self, sample_result):
        assert not sample_result.has_errors()

    def test_get_section_count(self, sample_result):
        assert sample_result.get_section_count() == 3

    def test_get_table_count(self, sample_result):
        assert sample_result.get_table_count() >= 1


class TestExcelEmptyFile:
    @pytest.fixture
    def empty_result(self, parser, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        f = tmp_path / "empty.xlsx"
        wb.save(str(f))
        wb.close()
        return parser.parse(f)

    def test_parse_empty(self, empty_result):
        assert isinstance(empty_result, DocumentResult)
        assert empty_result.source_format == "xlsx"
        assert len(empty_result.sections) == 1

    def test_no_errors_empty(self, empty_result):
        assert empty_result.errors == []


class TestExcelSingleSheet:
    @pytest.fixture
    def single_result(self, parser, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Value")
        ws.cell(row=2, column=1, value="Alpha")
        ws.cell(row=2, column=2, value=100)
        f = tmp_path / "single.xlsx"
        wb.save(str(f))
        wb.close()
        return parser.parse(f)

    def test_single_sheet(self, single_result):
        assert len(single_result.sections) == 1
        assert single_result.sections[0].title == "Data"

    def test_single_table(self, single_result):
        assert len(single_result.tables) == 1
        t = single_result.tables[0]
        assert t.headers == ["Name", "Value"]
        assert len(t.rows) == 1
        assert t.rows[0][0] == "Alpha"
        assert t.rows[0][1] == "100"


class TestExcelFormulaOnly:
    @pytest.fixture
    def formula_result(self, parser, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formulas"
        ws.cell(row=1, column=1, value="X")
        ws.cell(row=1, column=2, value="Y")
        ws.cell(row=1, column=3, value="Result")
        ws.cell(row=2, column=1, value=10)
        ws.cell(row=2, column=2, value=20)
        ws.cell(row=2, column=3).value = "=A2+B2"
        ws.cell(row=3, column=1, value=5)
        ws.cell(row=3, column=2, value=3)
        ws.cell(row=3, column=3).value = "=A3*B3"
        ws.cell(row=4, column=3).value = "=MAX(C2:C3)"
        f = tmp_path / "formulas.xlsx"
        wb.save(str(f))
        wb.close()
        return parser.parse(f)

    def test_formulas_extracted(self, formula_result):
        assert len(formula_result.formulas) == 3

    def test_formula_descriptions(self, formula_result):
        descs = [f.description for f in formula_result.formulas]
        assert any("Addition" in d for d in descs)
        assert any("Multiplication" in d for d in descs)
        assert any("MAX" in d or "maximum" in d.lower() for d in descs)

    def test_formula_deps_parsed(self, formula_result):
        add_f = [f for f in formula_result.formulas if "+" in f.formula_text][0]
        assert "A2" in add_f.dependencies
        assert "B2" in add_f.dependencies
