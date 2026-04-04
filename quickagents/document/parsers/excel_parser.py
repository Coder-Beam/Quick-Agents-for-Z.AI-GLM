"""
Excel Parser - Layer 1 parser for Excel files (.xlsx).

Uses openpyxl for cell data, formulas, chart extraction,
and requirement matrix/feature list structure recognition.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
import logging
import re

from . import BaseParser
from .utils import build_structure_tree_stack
from ..models import (
    DocumentResult,
    DocumentSection,
    DocumentTable,
    DocumentFormula,
)

logger = logging.getLogger(__name__)

_REQ_ID_PATTERNS = [
    re.compile(r"^(REQ|FR|NFR|UC|SRS|SDD|BR|AR|IR)-?\d+", re.IGNORECASE),
    re.compile(r"^\d{2,}[.\-\s]", re.IGNORECASE),
    re.compile(r"^(功能|需求|编号|序号|ID|No\.?)", re.IGNORECASE),
]

_MATRIX_HEADER_KEYWORDS = {
    "需求",
    "功能",
    "编号",
    "id",
    "no",
    "序号",
    "标题",
    "名称",
    "描述",
    "优先级",
    "状态",
    "模块",
    "负责人",
    "备注",
    "说明",
    "requirement",
    "feature",
    "priority",
    "status",
    "module",
    "description",
    "assignee",
    "remark",
    "title",
    "name",
}


class ExcelParser(BaseParser):
    """Excel parser using openpyxl"""

    SUPPORTED_FORMATS = ["xlsx", "xls"]
    REQUIRES_DEPENDENCIES = ["openpyxl"]
    PARSER_NAME = "excel"

    def __init__(self):
        super().__init__()
        self._openpyxl = None
        if self._deps_available:
            import openpyxl

            self._openpyxl = openpyxl

    def parse(self, file_path: Path) -> DocumentResult:
        if not self._deps_available:
            raise ImportError(
                "Excel parsing requires openpyxl. "
                "Install with: pip install quickagents[document]"
            )

        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        logger.info(f"Parsing Excel: {file_path}")
        metadata = self._extract_metadata(file_path)

        wb = self._openpyxl.load_workbook(
            str(file_path),
            data_only=False,
            read_only=True,
        )

        try:
            title = self._extract_title(wb, file_path)
            sheets_info = self._extract_sheets_overview(wb)
            sections = self._extract_sections(wb)
            paragraphs = self._extract_paragraphs(wb)
            tables = self._extract_tables(wb)
            formulas = self._extract_formulas(wb)
            structure_tree = build_structure_tree_stack(sections)

            return DocumentResult(
                source_file=str(file_path),
                source_format="xlsx",
                title=title,
                sections=sections,
                paragraphs=paragraphs,
                tables=tables,
                images=[],
                formulas=formulas,
                structure_tree=structure_tree,
                metadata={**metadata, **sheets_info},
                raw_text=self._extract_raw_text(wb),
                errors=[],
            )
        finally:
            wb.close()

    def _extract_title(self, wb, file_path: Path) -> str:
        if wb.properties and wb.properties.title:
            return wb.properties.title
        if len(wb.sheetnames) > 0:
            first = wb.sheetnames[0]
            if first.lower() not in ("sheet", "sheet1", "数据", "data"):
                return first
        return file_path.stem

    def _extract_sheets_overview(self, wb) -> Dict[str, Any]:
        info: Dict[str, Any] = {
            "sheet_count": len(wb.sheetnames),
            "sheet_names": list(wb.sheetnames),
        }
        props = wb.properties
        if props:
            if props.creator:
                info["creator"] = props.creator
            if props.created:
                info["created"] = props.created.isoformat()
            if props.modified:
                info["modified"] = props.modified.isoformat()
        return info

    def _extract_sections(self, wb) -> List[DocumentSection]:
        sections: List[DocumentSection] = []
        counter = 0
        for sheet_name in wb.sheetnames:
            counter += 1
            ws = wb[sheet_name]
            max_r = ws.max_row or 0
            max_c = ws.max_column or 0
            sections.append(
                DocumentSection(
                    section_id=f"S{counter:03d}",
                    title=sheet_name,
                    level=1,
                    page_number=counter,
                    content=f"Sheet with {max_r} rows x {max_c} columns",
                )
            )
        return sections

    def _extract_paragraphs(self, wb) -> List[str]:
        paragraphs: List[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(max_col=1):
                cell = row[0]
                if cell.value and isinstance(cell.value, str):
                    text = cell.value.strip()
                    if text and len(text) > 5:
                        paragraphs.append(text)
        return paragraphs

    def _extract_raw_text(self, wb) -> str:
        parts: List[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## {sheet_name}")
            for row in ws.iter_rows():
                cells = []
                for cell in row:
                    val = cell.value
                    if val is not None:
                        cells.append(str(val))
                if cells:
                    parts.append(" | ".join(cells))
        return "\n".join(parts)

    def _extract_tables(self, wb) -> List[DocumentTable]:
        tables: List[DocumentTable] = []
        table_counter = 0
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            max_r = ws.max_row or 0
            max_c = ws.max_column or 0
            if max_r < 2 or max_c < 2:
                continue

            matrix_info = self._detect_matrix_structure(ws, max_r, max_c)

            rows_data = self._read_all_rows(ws, max_r, max_c)

            if not rows_data:
                continue

            headers = rows_data[0]
            data_rows = rows_data[1:]

            caption = self._find_table_caption(ws, sheet_name, matrix_info)

            table_counter += 1
            tables.append(
                DocumentTable(
                    table_id=f"T{table_counter:03d}",
                    page_number=sheet_idx + 1,
                    caption=caption,
                    headers=headers,
                    rows=data_rows,
                )
            )

        return tables

    def _read_all_rows(self, ws, max_row: int, max_col: int) -> List[List[str]]:
        rows_data: List[List[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            cells = []
            for cell in row:
                val = cell.value
                if val is None:
                    cells.append("")
                elif isinstance(val, str):
                    cells.append(val.strip())
                else:
                    cells.append(str(val))
            rows_data.append(cells)
        return rows_data

    def _extract_formulas(self, wb) -> List[DocumentFormula]:
        formulas: List[DocumentFormula] = []
        counter = 0
        seen = set()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        cell_ref = cell.coordinate
                        if cell_ref in seen:
                            continue
                        seen.add(cell_ref)
                        counter += 1
                        deps = self._parse_formula_dependencies(val)
                        formulas.append(
                            DocumentFormula(
                                formula_id=f"F{counter:03d}",
                                formula_text=val,
                                description=self._describe_formula(val),
                                cell_ref=f"{sheet_name}!{cell_ref}",
                                dependencies=deps,
                                sheet_name=sheet_name,
                            )
                        )
        return formulas

    def _parse_formula_dependencies(self, formula: str) -> List[str]:
        refs = re.findall(
            r"(?:'?[^']*'!)?([A-Z]{1,3}\d{1,5}(?::[A-Z]{1,3}\d{1,5})?)",
            formula,
        )
        return refs[:20]

    def _describe_formula(self, formula: str) -> str:
        upper = formula.upper()
        funcs = re.findall(r"([A-Z]+)\s*\(", upper)
        if not funcs:
            if "*" in formula:
                return f"Multiplication: {formula[:80]}"
            if "+" in formula:
                return f"Addition: {formula[:80]}"
            if "-" in formula[1:]:
                return f"Subtraction: {formula[:80]}"
            if "/" in formula:
                return f"Division: {formula[:80]}"
            return f"Formula: {formula[:80]}"
        main = funcs[0]
        desc_map = {
            "SUM": "Sum calculation",
            "AVERAGE": "Average calculation",
            "COUNT": "Count values",
            "COUNTA": "Count non-empty values",
            "COUNTIF": "Conditional count",
            "COUNTIFS": "Multi-condition count",
            "IF": "Conditional logic",
            "IFS": "Multi-branch conditional",
            "VLOOKUP": "Vertical lookup",
            "HLOOKUP": "Horizontal lookup",
            "XLOOKUP": "Extended lookup",
            "INDEX": "Index reference",
            "MATCH": "Position match",
            "SUMIF": "Conditional sum",
            "SUMIFS": "Multi-condition sum",
            "AVERAGEIF": "Conditional average",
            "CONCATENATE": "Text concatenation",
            "CONCAT": "Text concatenation",
            "TEXT": "Text formatting",
            "LEFT": "Extract left characters",
            "RIGHT": "Extract right characters",
            "MID": "Extract middle characters",
            "LEN": "String length",
            "MAX": "Find maximum",
            "MIN": "Find minimum",
            "ROUND": "Round number",
            "ABS": "Absolute value",
            "TODAY": "Current date",
            "NOW": "Current datetime",
            "DATE": "Date construction",
            "YEAR": "Extract year",
            "MONTH": "Extract month",
            "DAY": "Extract day",
        }
        desc = desc_map.get(main, f"{main} function")
        return f"{desc}: {formula[:80]}"

    def _detect_matrix_structure(
        self, ws, max_row: int, max_col: int
    ) -> Dict[str, Any]:
        if max_row < 2 or max_col < 2:
            return {"is_matrix": False}

        header_row = []
        for col in range(1, min(max_col + 1, 20)):
            cell = ws.cell(row=1, column=col)
            val = cell.value
            if val:
                header_row.append(str(val).strip().lower())

        matrix_score = sum(
            1 for h in header_row if any(kw in h for kw in _MATRIX_HEADER_KEYWORDS)
        )

        has_req_ids = False
        for r in range(2, min(max_row + 1, 6)):
            first_cell = ws.cell(row=r, column=1).value
            if first_cell and self._looks_like_req_id(str(first_cell)):
                has_req_ids = True
                break

        is_matrix = matrix_score >= 2 or (matrix_score >= 1 and has_req_ids)

        return {
            "is_matrix": is_matrix,
            "matrix_score": matrix_score,
            "has_requirement_ids": has_req_ids,
        }

    def _looks_like_req_id(self, text: str) -> bool:
        text = text.strip()
        if not text:
            return False
        for pattern in _REQ_ID_PATTERNS:
            if pattern.match(text):
                return True
        return False

    def _find_table_caption(
        self, ws, sheet_name: str, matrix_info: Dict
    ) -> Optional[str]:
        if matrix_info.get("is_matrix"):
            return f"Requirement matrix: {sheet_name}"

        for col in range(1, min((ws.max_column or 0) + 1, 5)):
            val = ws.cell(row=1, column=col).value
            if val and isinstance(val, str):
                lower = val.strip().lower()
                if any(
                    kw in lower
                    for kw in ("表", "table", "list", "列表", "矩阵", "matrix")
                ):
                    return val.strip()
        return None
